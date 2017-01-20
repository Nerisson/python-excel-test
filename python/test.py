from openpyxl.styles import PatternFill
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()


for i in range(11):
	v = int((213-38)/10*i+38)
	color = '%02x%02x%02x' % (130, v, 255)
	fill = PatternFill(start_color=color, fill_type='solid')
	ws['A'+str(4+i)].fill = fill

# Save the file
wb.save("sample.xlsx")

